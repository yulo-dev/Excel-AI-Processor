# backend/app.py

import os
from flask import Flask, request, jsonify, Response
from flask_cors import CORS
import pandas as pd
from dotenv import load_dotenv
import google.generativeai as genai
import logging
import io
import sys
import json
import numpy as np
import re
import datetime
from openpyxl import load_workbook, Workbook # Import Workbook for new sheets

# Load environment variables from .env file
load_dotenv()

# --- Gemini API Configuration ---
API_KEY = os.getenv("GOOGLE_API_KEY")
if not API_KEY:
    print("ERROR: GOOGLE_API_KEY is not set. Please provide your Gemini API key in the .env file.")
else:
    # Suppress verbose Gemini API logging to keep console clean
    logging.getLogger('google.generativeai').addHandler(logging.NullHandler())
    logging.getLogger('google.generativeai').propagate = False
    genai.configure(api_key=API_KEY)


# Initialize Flask app
app = Flask(__name__)
CORS(app) # Enable CORS for frontend communication

# Configure upload and download folders
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
DOWNLOAD_FOLDER = os.path.join(BASE_DIR, 'downloads')

# Create folders if they don't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(DOWNLOAD_FOLDER):
    os.makedirs(DOWNLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limit max file size to 16MB

# Allowed file extensions for Excel files
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """Checks if the uploaded file has an allowed extension."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Core change: parse_ai_response_to_list now strictly parses Markdown lists
# AI will always return Markdown lists, so this function no longer needs to check is_markdown_list_expected
def parse_ai_response_to_list(ai_response_text: str) -> list:
    """
    Parses AI text response to strictly extract items from a Markdown list.
    This version aims to precisely extract list content, filtering out any explanatory text or unrelated symbols.
    """
    lines = ai_response_text.split('\n')
    parsed_items = []

    # Define common explanatory or non-content line starting keywords
    filter_keywords = [
        "response:", "here is", "list below", "results:", "pure result list",
        "could not generate a valid response", "no valid command", "error", "hint", "summary",
        "this is about", "here's what you asked for", "based on your command", "okay", "please refer",
        "provided for you below", "here is the de-duplicated list", "these are the extracted company names",
        "here are the differences", "difference list below",
        "hello", "this is your", "as per your request", "certainly" # Added more general opening phrases
    ]

    for line in lines:
        stripped_line = line.strip()
        if not stripped_line: # Ignore empty lines
            continue

        # Check if it's an explanatory or non-content line (these lines should be completely ignored)
        # Avoid deleting actual list items (e.g., if an item itself is "results:")
        is_potential_explanatory_line = any(keyword in stripped_line.lower() for keyword in filter_keywords)

        # If it's an explanatory line and does not appear to be an actual list item, skip
        if is_potential_explanatory_line and not stripped_line.startswith('*'):
            continue # Only skip if it's an explanatory line and not a markdown list item

        # Strictly match lines starting with a Markdown list symbol, extract content and clean it
        # Must start with *, followed by a space, then the content
        match = re.match(r'^\*\s*(.+)$', stripped_line)
        if match:
            item = match.group(1).strip()
            # Remove potential bold formatting (**)
            item = re.sub(r'^\**', '', item)
            item = re.sub(r'\**$', '', item)
            # Remove parenthesized data after company names (e.g., (4453))
            item = re.sub(r'\s*\(\d+\)$', '', item).strip()
            # Remove potential numbering (e.g., "1. Item", "2) Item")
            item = re.sub(r'^\d+\.?\s*\)?\s*', '', item).strip()

            if item:
                parsed_items.append(item)

    return parsed_items


# Core change: process_ai_command now always instructs AI to output a Markdown list
def process_ai_command(dataframe: pd.DataFrame, command: str) -> str:
    """
    Processes AI commands using the Google Gemini model.
    AI will be instructed to always output a Markdown list for consistency.
    """
    if not API_KEY:
        return "ERROR: Gemini API is not configured, cannot process AI command."

    old_stdout = sys.stdout
    redirected_output = io.StringIO()
    sys.stdout = redirected_output

    try:
        df_csv_string = dataframe.to_csv(index=False)
        original_rows_count = dataframe.shape[0]

        # Construct column mapping reference, precise and emphatic
        column_names_list = dataframe.columns.tolist()
        column_info_str = "Please strictly refer to the following actual column names (Column Names) in the Excel data:\n"
        for i, col_name in enumerate(column_names_list):
            column_info_str += f"- Actual Column Name: '{col_name}' (May correspond to traditional Excel Column {chr(65 + i)})\n"
        column_info_str += "\n**EXTREMELY IMPORTANT:** When a user command mentions an Excel column letter (e.g., 'Column C', 'Column E'), you **MUST** accurately determine and use the actual column name from the CSV data provided (e.g., 'Unnamed: 2' for 'Column C').\n"
        column_info_str += "For example, if the user mentions 'Column C', and its corresponding actual column name is 'Unnamed: 2', you **MUST USE** 'Unnamed: 2' to process the data.\n"
        column_info_str += "Ensure you only process columns explicitly specified by the user, do not extend to other unmentioned columns.\n"
        column_info_str += "If the command involves de-duplication, please strictly perform the de-duplication operation.\n" # Emphasize de-duplication


        prompt = f"""You are a powerful data analysis assistant.
You will receive a segment of data from an Excel file (in CSV format), and a command from the user regarding this data.
Please analyze the provided data based on the user's command and give a clear, concise response.

{column_info_str}

---
Excel Data (CSV Format):
{df_csv_string}
---

User Command: {command}
---

**EXTREMELY IMPORTANT INSTRUCTION: Please output all extracted or calculated results in Markdown list format.**
**This is the ONLY expected output format. Please strictly adhere to all the following requirements:**
1.  **ONLY provide a Markdown list.**
2.  **Each list item MUST start with `* ` (an asterisk followed by a space).**
3.  **Each list item MUST be on a separate line, with no extra line breaks, empty lines, or symbols.**
4.  **ABSOLUTELY DO NOT include any additional text, explanations, headings, summaries, introductions, greetings, or any non-list content.**
5.  Ensure each list item is a concise, single result, without extra parentheses, numbers, Markdown bolding (e.g., **Name**), or redundant descriptions.
6.  If the command requires generating a new column value for each row of the original data, ensure the output list has the same length as the original data's row count ({original_rows_count} rows), and provide a value for each row (output blank or N/A if a value cannot be generated).

Your Response (ONLY a Markdown list, with no other text, formatting, or explanations):
"""

        print(f"\n--- Full Prompt Sent to Gemini ---\n{prompt}\n--- End of Prompt ---") # Log: Print full prompt

        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)

        if response and response.candidates and response.candidates[0].content and response.candidates[0].content.parts and response.candidates[0].content.parts[0].text:
            return response.candidates[0].content.parts[0].text
        else:
            return "Gemini model could not generate a valid response. Please try another command."

    except Exception as e:
        print(f"Error calling Gemini API: {e}")
        return f"An error occurred while processing AI command: {str(e)}"
    finally:
        sys.stdout = old_stdout # Ensure logs are visible even on error


@app.route('/')
def home():
    """Home route or health check."""
    return jsonify({"message": "Welcome to the Excel AI Processor backend!"}), 200

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handles Excel file upload and AI command processing."""
    print("Received file upload and processing request...")

    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request."}), 400

    file = request.files['file']
    command = request.form.get('command', '').strip()
    sheet_name = request.form.get('sheet_name', '').strip()
    output_mode = request.form.get('output_mode', 'new_sheet_original_file')
    new_column_name = request.form.get('new_column_name', '').strip()

    print(f"DEBUG: Backend received command from frontend: '{command}', Output Mode: {output_mode}, New Column Name: '{new_column_name}'")

    if file.filename == '':
        return jsonify({"error": "No file selected."}), 400

    if not command:
        return jsonify({"error": "Please provide an AI command!"}), 400

    if output_mode == 'new_column_original_sheet' and not new_column_name:
        return jsonify({"error": "When adding a new column, please specify its name!"}), 400

    if file and allowed_file(file.filename):
        original_filename = file.filename
        filepath_original = os.path.join(app.config['UPLOAD_FOLDER'], original_filename)

        # Ensure the file is closed before attempting to remove it (for overwriting)
        if os.path.exists(filepath_original):
            try:
                os.remove(filepath_original)
                print(f"DEBUG: Removed old file {filepath_original}.")
            except Exception as e:
                print(f"WARNING: Could not remove old file {filepath_original}: {e}. It might be in use.")
                return jsonify({"error": f"Could not process file. Please ensure the Excel file is not open or in use, and re-upload. Detailed error: {e}"}), 400

        file.save(filepath_original) # Save the original file
        print(f"Original file saved to: {filepath_original}")

        try:
            xls = pd.ExcelFile(filepath_original)
            if sheet_name and sheet_name in xls.sheet_names:
                actual_sheet_name_read = sheet_name
                df = pd.read_excel(xls, sheet_name=actual_sheet_name_read, header=0)
                print(f"Successfully read sheet '{actual_sheet_name_read}'. First 5 rows of data:")
            else:
                actual_sheet_name_read = xls.sheet_names[0]
                df = pd.read_excel(xls, sheet_name=actual_sheet_name_read, header=0)
                print(f"No sheet name specified. Defaulting to read the first sheet '{actual_sheet_name_read}'. First 5 rows of data:")

            # --- New Log: Display actual column names read by Pandas ---
            print(f"\n--- Actual Column Names Read by Pandas ---\n{df.columns.tolist()}\n--- End of Actual Column Names ---")
            # -----------------------------------------------------------

            original_df_rows = df.shape[0]
            df = df.replace({np.nan: None}) # Replace NaN with None for consistent AI input

            print(df.head())

            # Core change: AI is always instructed to output a Markdown list.
            # parse_ai_response_to_list will now always expect a Markdown list.
            ai_response_text = process_ai_command(df, command) # No longer pass output_mode
            parsed_results = parse_ai_response_to_list(ai_response_text) # No longer pass is_markdown_list_expected

            print(f"\n--- Raw AI Response ---\n{ai_response_text}\n--- End of Raw AI Response ---") # Prominently print raw AI response
            print(f"DEBUG: Parsed AI list content: {parsed_results}") # Log parsed AI results
            print(f"DEBUG: Parsed AI list length: {len(parsed_results)}, Original data rows: {original_df_rows}") # Log lengths

            download_url = None

            if output_mode == "new_column_original_sheet":
                # Ensure AI output list length is compatible with DataFrame index
                # If AI is expected to generate a value for each row, but the number of results don't match, issue a warning
                if len(parsed_results) != original_df_rows:
                    warning_msg = f"WARNING: AI output count ({len(parsed_results)}) does not match original data rows ({original_df_rows}). Results will be adjusted to fit the new column (may be truncated or padded with blanks)."
                    print(warning_msg)
                    ai_response_text += f"\n\nNote: {warning_msg}"

                # Critical final fix: Create an empty Series with the same length as the original DataFrame, then populate with AI results
                new_column_series = pd.Series(index=df.index, dtype=object) # Create an empty Series with the correct index
                # Populate the beginning of the new Series with parsed_results
                # If parsed_results is shorter, the rest will be NaN (blank in Excel)
                # If parsed_results is longer, it will be truncated
                new_column_series.iloc[:min(len(parsed_results), original_df_rows)] = parsed_results[:min(len(parsed_results), original_df_rows)]
                df[new_column_name] = new_column_series # Assign to the new column in DataFrame

                print(f"DEBUG: New column '{new_column_name}' added to DataFrame.")
                print(f"DEBUG: DataFrame tail (with new column): \n{df.tail()}") # Log DataFrame state

                try:
                    book = load_workbook(filepath_original)

                    if actual_sheet_name_read in book.sheetnames:
                        old_sheet_index = book.sheetnames.index(actual_sheet_name_read)
                        book.remove(book[actual_sheet_name_read])
                        print(f"DEBUG: Old sheet '{actual_sheet_name_read}' removed from Excel workbook.")

                    temp_excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(temp_excel_buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=actual_sheet_name_read, index=False)
                    temp_excel_buffer.seek(0)

                    new_sheet_book = load_workbook(temp_excel_buffer)
                    new_sheet = new_sheet_book.active

                    book.add_sheet(new_sheet, index=old_sheet_index if 'old_sheet_index' in locals() else 0)
                    new_sheet.title = actual_sheet_name_read # Ensure the sheet name is preserved

                    book.save(filepath_original) # Save the modified workbook
                    print(f"AI processed results added as new column '{new_column_name}' to sheet '{actual_sheet_name_read}' in original file '{original_filename}', preserving other sheets.")

                    host_url = request.host_url.rstrip('/')
                    download_url = f"{host_url}/download/{original_filename}"
                    print(f"DEBUG: Generated download URL (pointing to modified original file - new column mode): {download_url}")

                except Exception as writer_e:
                    print(f"ERROR: Problem writing new column results to original Excel file: {writer_e}")
                    ai_response_text += f"\n\nERROR: Could not write results as a new column to the original Excel file. Please ensure the file is not in use and re-upload. Detailed error: {writer_e}"
                    download_url = None
                    print(f"DEBUG: Failed to write to original Excel file, no download link generated (new column mode). Error: {writer_e}")

            elif output_mode == "new_sheet_original_file":
                if parsed_results:
                    output_df = pd.DataFrame(parsed_results, columns=['AI_Processed_Result']) # Column name for the new sheet
                    modified_excel_path = filepath_original
                    try:
                        with pd.ExcelWriter(modified_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            output_df.to_excel(writer, sheet_name="AI_Results", index=False) # Fixed sheet name
                            print(f"AI processed results written to 'AI_Results' sheet in original file '{original_filename}'.")

                        host_url = request.host_url.rstrip('/')
                        download_url = f"{host_url}/download/{original_filename}"
                        print(f"DEBUG: Generated download URL (pointing to modified original file - new sheet mode): {download_url}")
                    except Exception as writer_e:
                        print(f"ERROR: Problem writing results to original Excel file (new sheet mode): {writer_e}")
                        ai_response_text += f"\n\nERROR: Could not write results to original Excel file. Please ensure the file is not in use and re-upload. Detailed error: {writer_e}"
                        download_url = None
                        print(f"DEBUG: Failed to write to original Excel file, no download link generated (new sheet mode). Error: {writer_e}")
                else:
                    print("DEBUG: AI response could not be parsed into a list, no downloadable Excel file generated (new sheet mode).")
                    ai_response_text += "\n\nNote: AI response could not be parsed into a list, unable to generate result Excel file."
                    download_url = None

            elif output_mode == "new_excel_file":
                if parsed_results:
                    output_df = pd.DataFrame(parsed_results, columns=['AI_Processed_Result']) # Column name for the new file
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_filename_new = f"ai_result_{timestamp}.xlsx"
                    output_filepath_new = os.path.join(app.config['DOWNLOAD_FOLDER'], output_filename_new)

                    output_df.to_excel(output_filepath_new, index=False)
                    print(f"AI processed results saved as new Excel file: {output_filepath_new}")

                    host_url = request.host_url.rstrip('/')
                    download_url = f"{host_url}/download_new/{output_filename_new}"
                    print(f"DEBUG: Generated download URL (pointing to new file): {download_url}")
                else:
                    print("DEBUG: AI response could not be parsed into a list, no downloadable Excel file generated (new file mode).")
                    ai_response_text += "\n\nNote: AI response could not be parsed into a list, unable to generate result Excel file."
                    download_url = None

            else: # If output_mode is unknown
                print(f"ERROR: Unknown output mode: {output_mode}")
                ai_response_text += "\n\nERROR: Unknown output mode, unable to process download request."
                download_url = None

            response_data = {
                "message": "File upload, reading, and AI processing successful!",
                "filename": original_filename,
                "data_preview": df.head().to_dict(orient='records'),
                "ai_response": ai_response_text,
                "download_url": download_url
            }

            # Redirect stdout back to original for JSON serialization logging
            temp_stdout = sys.stdout
            sys.stdout = io.StringIO()

            json_string = json.dumps(response_data, ensure_ascii=False)

            sys.stdout = temp_stdout # Restore original stdout

            return Response(json_string, mimetype='application/json')

        except Exception as e:
            print(f"ERROR: An error occurred while processing file or AI command: {e}")
            if "No sheet named" in str(e):
                return jsonify({"error": f"Could not find sheet named '{sheet_name}'. Please check the name."}), 400
            elif "Worksheet named" in str(e) and "not found" in str(e):
                 return jsonify({"error": f"Could not find sheet named '{sheet_name}'. Please check the name."}), 400
            else:
                return jsonify({"error": f"Processing failed: {str(e)}"}), 500
    else:
        return jsonify({"error": "File type not allowed. Only .xlsx and .xls files are accepted."}), 400

@app.route('/download/<filename>')
def download_original_file(filename):
    """Serves the modified original file for download."""
    print(f"DEBUG: Received request to download original file: {filename}")
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not os.path.exists(full_path):
        print(f"ERROR: File {full_path} not found in the specified directory.")
        return jsonify({"error": "File not found."}), 404

    try:
        with open(full_path, 'rb') as f:
            file_data = f.read()

        response = Response(file_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response
    except Exception as e:
        print(f"ERROR: Problem reading and serving original file {filename}: {e}")
        return jsonify({"error": f"Failed to download file: {e}"}), 500

@app.route('/download_new/<filename>')
def download_new_file(filename):
    """Serves the newly created file for download."""
    print(f"DEBUG: Received request to download new file: {filename}")
    full_path = os.path.join(app.config['DOWNLOAD_FOLDER'], filename)

    if not os.path.exists(full_path):
        print(f"ERROR: File {full_path} not found in the specified directory.")
        return jsonify({"error": "File not found."}), 404

    try:
        with open(full_path, 'rb') as f:
            file_data = f.read()

        response = Response(file_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response
    except Exception as e:
        print(f"ERROR: Problem reading and serving new file {filename}: {e}")
        return jsonify({"error": f"Failed to download file: {e}"}), 500


if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
